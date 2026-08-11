"""Pruebas del reporte Excel: estructura, formatos y color por precedencia."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from reconciliacion.dominio.enums import Clasificacion, TipoFraude
from reconciliacion.dominio.modelos import Retencion
from reconciliacion.exportadores.excel import (
    COLUMNAS,
    RELLENO_FRAUDE,
    RELLENO_HALLAZGO,
    RELLENO_RECONCILIADO,
    ExportadorExcel,
)

TITULOS_ESPERADOS = [
    "ID_Transaccion", "Clasificacion", "Presente_CSV", "Presente_SQLite", "Presente_JSON",
    "Monto_CSV", "Monto_SQLite", "Monto_JSON", "Diferencia_Monto",
    "Fecha_CSV", "Fecha_SQLite", "Fecha_JSON",
    "Estado_CSV", "Estado_SQLite", "Estado_JSON", "Banco",
    "Marca_CSV", "Marca_SQLite", "Marca_JSON", "Marca_Coincide",
    "Retencion_IVA_ICA", "Retencion_Fuente_IVA", "Retencion_IVA_ICA_Fuente",
    "Centro_Costo", "ID_Movimiento_Bancario",
    "Es_Fraude", "Tipo_Fraude", "Nivel_Riesgo", "Observaciones",
]


@pytest.fixture
def hoja_generada(tmp_path: Path, fabricar_transaccion):
    """Genera un Excel con un caso de cada tipo y devuelve su hoja."""
    reconciliada = fabricar_transaccion(id_transaccion="TRX0001")
    reconciliada.clasificaciones = [Clasificacion.RECONCILIADO]

    con_hallazgo = fabricar_transaccion(id_transaccion="TRX0002", en_json=False)
    con_hallazgo.clasificaciones = [Clasificacion.NO_ENCONTRADO_EN_JSON]

    con_fraude = fabricar_transaccion(id_transaccion="TRX0003")
    con_fraude.clasificaciones = [Clasificacion.RECONCILIADO]
    con_fraude.tipos_fraude = [TipoFraude.MONTO]

    ruta = ExportadorExcel().exportar(
        [reconciliada, con_hallazgo, con_fraude], ruta=tmp_path / "reporte.xlsx"
    )
    return load_workbook(ruta)


class TestEstructura:
    def test_una_sola_hoja_llamada_reconciliacion(self, hoja_generada) -> None:
        assert hoja_generada.sheetnames == ["Reconciliacion"]

    def test_veintinueve_columnas_en_el_orden_pedido(self, hoja_generada) -> None:
        hoja = hoja_generada["Reconciliacion"]
        assert [c.value for c in hoja[1]] == TITULOS_ESPERADOS
        assert len(COLUMNAS) == 29

    def test_la_tabla_empieza_en_a1(self, hoja_generada) -> None:
        assert hoja_generada["Reconciliacion"]["A1"].value == "ID_Transaccion"

    def test_una_fila_por_transaccion(self, hoja_generada) -> None:
        assert hoja_generada["Reconciliacion"].max_row == 4  # encabezado + 3

    def test_encabezado_congelado_y_con_autofiltro(self, hoja_generada) -> None:
        hoja = hoja_generada["Reconciliacion"]
        assert hoja.freeze_panes == "A2"
        assert hoja.auto_filter.ref == "A1:AC4"

    def test_encabezado_en_negrita_y_con_fondo(self, hoja_generada) -> None:
        celda = hoja_generada["Reconciliacion"]["A1"]
        assert celda.font.bold is True
        assert celda.fill.fgColor.rgb.endswith("305496")


class TestFormatos:
    def test_los_montos_son_numeros_con_formato_de_moneda(self, hoja_generada) -> None:
        celda = hoja_generada["Reconciliacion"].cell(row=2, column=6)
        assert isinstance(celda.value, (int, float))
        assert "$" in celda.number_format

    def test_las_fechas_son_fechas_y_no_texto(self, hoja_generada) -> None:
        celda = hoja_generada["Reconciliacion"].cell(row=2, column=10)
        assert isinstance(celda.value, datetime)
        assert "DD/MM/YYYY" in celda.number_format

    def test_las_columnas_tienen_ancho_asignado(self, hoja_generada) -> None:
        hoja = hoja_generada["Reconciliacion"]
        assert hoja.column_dimensions["A"].width > 0


class TestColorPorPrecedencia:
    """Naranja si hay fraude; si no rojo ante hallazgo; si no verde."""

    def test_reconciliada_en_verde(self, hoja_generada) -> None:
        celda = hoja_generada["Reconciliacion"].cell(row=2, column=1)
        assert celda.fill.fgColor.rgb == RELLENO_RECONCILIADO.fgColor.rgb

    def test_hallazgo_en_rojo(self, hoja_generada) -> None:
        celda = hoja_generada["Reconciliacion"].cell(row=3, column=1)
        assert celda.fill.fgColor.rgb == RELLENO_HALLAZGO.fgColor.rgb

    def test_el_fraude_manda_sobre_el_verde(self, hoja_generada) -> None:
        """TRX0003 esta RECONCILIADO pero es fraude: debe salir naranja."""
        hoja = hoja_generada["Reconciliacion"]
        assert hoja.cell(row=4, column=2).value == "RECONCILIADO"
        assert hoja.cell(row=4, column=26).value == "SI"
        assert hoja.cell(row=4, column=1).fill.fgColor.rgb == RELLENO_FRAUDE.fgColor.rgb


class TestContenidoDeLasCeldas:
    def test_presencia_como_si_o_no(self, hoja_generada) -> None:
        hoja = hoja_generada["Reconciliacion"]
        assert hoja.cell(row=2, column=3).value == "SI"
        assert hoja.cell(row=3, column=5).value == "NO"  # TRX0002 no esta en JSON

    def test_las_columnas_de_la_fuente_ausente_quedan_vacias(self, hoja_generada) -> None:
        """openpyxl guarda la cadena vacia como celda sin valor; ambas cuentan."""
        hoja = hoja_generada["Reconciliacion"]
        assert not hoja.cell(row=3, column=8).value  # Monto_JSON
        assert not hoja.cell(row=3, column=25).value  # ID_Movimiento_Bancario
        assert not hoja.cell(row=3, column=12).value  # Fecha_JSON

    def test_las_etiquetas_multiples_van_separadas_por_punto_y_coma(
        self, tmp_path: Path, fabricar_transaccion
    ) -> None:
        transaccion = fabricar_transaccion(id_transaccion="TRX0004", en_json=False)
        transaccion.clasificaciones = [
            Clasificacion.NO_ENCONTRADO_EN_JSON,
            Clasificacion.DISCREPANCIA_MONTO,
        ]
        transaccion.tipos_fraude = [TipoFraude.HORA, TipoFraude.PATRON]

        ruta = ExportadorExcel().exportar([transaccion], ruta=tmp_path / "multi.xlsx")
        hoja = load_workbook(ruta)["Reconciliacion"]
        assert hoja.cell(row=2, column=2).value == "NO_ENCONTRADO_EN_JSON;DISCREPANCIA_MONTO"
        assert hoja.cell(row=2, column=27).value == "FRAUDE_HORA;FRAUDE_PATRON"

    def test_las_retenciones_se_calculan_desde_el_csv(
        self, tmp_path: Path, fabricar_transaccion
    ) -> None:
        transaccion = fabricar_transaccion(
            csv_retenciones=[Retencion("iva", -100.0), Retencion("ica", -50.0)]
        )
        ruta = ExportadorExcel().exportar([transaccion], ruta=tmp_path / "ret.xlsx")
        hoja = load_workbook(ruta)["Reconciliacion"]
        assert hoja.cell(row=2, column=21).value == -150.0  # iva + ica
        assert hoja.cell(row=2, column=22).value == -100.0  # fuente + iva

    def test_sin_csv_las_retenciones_quedan_vacias(
        self, tmp_path: Path, fabricar_transaccion
    ) -> None:
        transaccion = fabricar_transaccion(en_csv=False, en_sqlite=False)
        ruta = ExportadorExcel().exportar([transaccion], ruta=tmp_path / "sin.xlsx")
        hoja = load_workbook(ruta)["Reconciliacion"]
        assert hoja.cell(row=2, column=21).value is None
