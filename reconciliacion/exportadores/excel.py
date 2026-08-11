"""El Excel, que es el entregable principal.

Una sola hoja, tabla desde A1, una fila por transaccion y las tres fuentes lado a
lado.

Lo unico que quiero destacar del diseno: las 29 columnas estan declaradas como
**datos** —una lista donde cada columna sabe su titulo, de donde sale su valor y
con que formato se muestra— y no como 29 bloques de codigo repetido. Escribir la
hoja es el mismo bucle sin importar cuantas columnas haya, y mover una es editar
una linea.

Los montos y las fechas se escriben como numero y fecha de verdad, no como texto,
para que contabilidad pueda filtrar, sumar y ordenar sin convertir nada a mano.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from reconciliacion.config import rutas
from reconciliacion.dominio.transaccion import TransaccionReconciliada
from reconciliacion.errores import ErrorExportacion
from reconciliacion.limpieza.marcas import marcas_coinciden
from reconciliacion.log import obtener_logger
from reconciliacion.progreso import ProgresoParcial, notificar

_log = obtener_logger(__name__)

NOMBRE_HOJA = "Reconciliacion"

FORMATO_MONEDA = '"$"#,##0'
FORMATO_MONEDA_DECIMAL = '"$"#,##0.00'
FORMATO_FECHA = "DD/MM/YYYY HH:MM"

# Paleta: encabezado azul oscuro y filas segun el hallazgo. Se usan los tonos
# convencionales de Excel para "malo" y "bueno" porque son los que un area
# contable ya reconoce sin necesidad de leyenda.
RELLENO_ENCABEZADO = PatternFill("solid", fgColor="305496")
RELLENO_FRAUDE = PatternFill("solid", fgColor="FFCC99")  # naranja
RELLENO_HALLAZGO = PatternFill("solid", fgColor="FFC7CE")  # rojo
RELLENO_RECONCILIADO = PatternFill("solid", fgColor="C6EFCE")  # verde

FUENTE_ENCABEZADO = Font(bold=True, color="FFFFFF")

ANCHO_MINIMO = 10
ANCHO_MAXIMO = 46


def _si_no(valor: bool) -> str:
    """Convierte un booleano al `SI`/`NO` que espera el reporte."""
    return "SI" if valor else "NO"


def _marca_coincide(transaccion: TransaccionReconciliada) -> str:
    """Devuelve `SI`/`NO` segun si las marcas de las fuentes presentes coinciden.

    Args:
        transaccion: Transaccion a evaluar.

    Returns:
        `SI` o `NO`, o cadena vacia si ninguna fuente aporta marca y por tanto
        no hay nada que comparar.
    """
    resultado = marcas_coinciden(
        transaccion.marca_csv, transaccion.marca_sqlite, transaccion.marca_json
    )
    if resultado is None:
        return ""
    return _si_no(resultado)


def _retencion(transaccion: TransaccionReconciliada, *entidades: str) -> Optional[float]:
    """Suma las retenciones pedidas, o vacio si la transaccion no esta en el CSV.

    Args:
        transaccion: Transaccion a evaluar.
        *entidades: Entidades a sumar.

    Returns:
        La suma (0 si las entidades no estan en la fila), o `None` si la
        transaccion no existe en el CSV y por tanto no tiene retenciones.
    """
    if transaccion.autorizacion is None:
        return None
    return transaccion.autorizacion.total_retencion(*entidades)


@dataclass(frozen=True)
class ColumnaReporte:
    """Definicion de una columna de la hoja.

    Attributes:
        titulo: Encabezado que se escribe en la fila 1.
        valor: Funcion que extrae el valor de una transaccion.
        formato: Formato de numero de Excel, o `None` para texto.
    """

    titulo: str
    valor: Callable[[TransaccionReconciliada], object]
    formato: Optional[str] = None


#: Las 29 columnas del reporte, en el orden exigido.
COLUMNAS: tuple[ColumnaReporte, ...] = (
    ColumnaReporte("ID_Transaccion", lambda t: t.id_transaccion),
    ColumnaReporte("Clasificacion", lambda t: ";".join(str(c) for c in t.clasificaciones)),
    ColumnaReporte("Presente_CSV", lambda t: _si_no(t.presente_csv)),
    ColumnaReporte("Presente_SQLite", lambda t: _si_no(t.presente_sqlite)),
    ColumnaReporte("Presente_JSON", lambda t: _si_no(t.presente_json)),
    ColumnaReporte("Monto_CSV", lambda t: t.monto_csv, FORMATO_MONEDA),
    ColumnaReporte("Monto_SQLite", lambda t: t.monto_sqlite, FORMATO_MONEDA),
    ColumnaReporte("Monto_JSON", lambda t: t.monto_json, FORMATO_MONEDA),
    ColumnaReporte("Diferencia_Monto", lambda t: t.diferencia_monto, FORMATO_MONEDA),
    ColumnaReporte("Fecha_CSV", lambda t: t.fecha_csv, FORMATO_FECHA),
    ColumnaReporte("Fecha_SQLite", lambda t: t.fecha_sqlite, FORMATO_FECHA),
    ColumnaReporte("Fecha_JSON", lambda t: t.fecha_json, FORMATO_FECHA),
    ColumnaReporte("Estado_CSV", lambda t: t.estado_csv or ""),
    ColumnaReporte("Estado_SQLite", lambda t: t.estado_sqlite or ""),
    ColumnaReporte("Estado_JSON", lambda t: t.estado_json or ""),
    ColumnaReporte("Banco", lambda t: t.banco or ""),
    ColumnaReporte("Marca_CSV", lambda t: t.marca_csv or ""),
    ColumnaReporte("Marca_SQLite", lambda t: t.marca_sqlite or ""),
    ColumnaReporte("Marca_JSON", lambda t: t.marca_json or ""),
    ColumnaReporte("Marca_Coincide", _marca_coincide),
    ColumnaReporte(
        "Retencion_IVA_ICA", lambda t: _retencion(t, "iva", "ica"), FORMATO_MONEDA_DECIMAL
    ),
    ColumnaReporte(
        "Retencion_Fuente_IVA", lambda t: _retencion(t, "fuente", "iva"), FORMATO_MONEDA_DECIMAL
    ),
    ColumnaReporte(
        "Retencion_IVA_ICA_Fuente",
        lambda t: _retencion(t, "iva", "ica", "fuente"),
        FORMATO_MONEDA_DECIMAL,
    ),
    ColumnaReporte(
        "Centro_Costo",
        lambda t: t.contabilizacion.centro_costo if t.contabilizacion else "",
    ),
    ColumnaReporte(
        "ID_Movimiento_Bancario",
        lambda t: t.movimiento.id_movimiento if t.movimiento else "",
    ),
    ColumnaReporte("Es_Fraude", lambda t: _si_no(t.es_fraude)),
    ColumnaReporte("Tipo_Fraude", lambda t: ";".join(str(f) for f in t.tipos_fraude)),
    ColumnaReporte("Nivel_Riesgo", lambda t: str(t.nivel_riesgo) if t.nivel_riesgo else ""),
    ColumnaReporte("Observaciones", lambda t: " ".join(t.observaciones)),
)


class ExportadorExcel:
    """Escribe el reporte de reconciliacion en un archivo .xlsx.

    Attributes:
        columnas: Definicion de las columnas a exportar.
    """

    def __init__(self, columnas: Sequence[ColumnaReporte] | None = None) -> None:
        """Crea el exportador.

        Args:
            columnas: Columnas a escribir. Si se omite se usan las 29 del
                enunciado.
        """
        self.columnas = tuple(columnas) if columnas is not None else COLUMNAS

    def exportar(
        self,
        transacciones: Sequence[TransaccionReconciliada],
        ruta: Path | None = None,
        progreso: Optional[ProgresoParcial] = None,
    ) -> Path:
        """Genera el archivo Excel con el detalle transaccion por transaccion.

        Args:
            transacciones: Universo reconciliado y evaluado.
            ruta: Destino del archivo. Si se omite se usa el configurado en
                `rutas`.
            progreso: Aviso opcional de avance parcial.

        Returns:
            La ruta del archivo escrito.

        Raises:
            ErrorExportacion: Si el archivo esta abierto o la carpeta destino
                no es escribible.
        """
        destino = Path(ruta) if ruta else rutas.RUTA_REPORTE_EXCEL
        rutas.asegurar_directorio_salida()

        libro = Workbook()
        hoja = libro.active
        hoja.title = NOMBRE_HOJA

        self._escribir_encabezados(hoja)
        for indice, transaccion in enumerate(transacciones, start=2):
            self._escribir_fila(hoja, indice, transaccion)
            notificar(progreso, indice - 1, len(transacciones))

        self._dar_formato(hoja, filas=len(transacciones) + 1)

        # Guardar es una sola operacion larga que no se puede trocear, asi que
        # se avisa antes de entrar en ella para que la interfaz no parezca
        # detenida en el ultimo tramo.
        _log.info("Guardando el archivo en disco...")

        try:
            libro.save(destino)
        except PermissionError as exc:
            raise ErrorExportacion(
                "No se pudo guardar el reporte porque el archivo esta abierto. "
                "Cierralo en Excel y vuelve a ejecutar.",
                detalle=str(exc),
            ) from exc
        except OSError as exc:
            raise ErrorExportacion(
                "No se pudo escribir el reporte en la carpeta de salida.",
                detalle=str(exc),
            ) from exc

        _log.info("Reporte generado: %s (%d filas)", destino, len(transacciones))
        return destino

    def _escribir_encabezados(self, hoja: Worksheet) -> None:
        """Escribe la fila 1 con los titulos de las columnas."""
        for numero, columna in enumerate(self.columnas, start=1):
            celda = hoja.cell(row=1, column=numero, value=columna.titulo)
            celda.font = FUENTE_ENCABEZADO
            celda.fill = RELLENO_ENCABEZADO
            celda.alignment = Alignment(horizontal="center", vertical="center")

    def _escribir_fila(
        self, hoja: Worksheet, fila: int, transaccion: TransaccionReconciliada
    ) -> None:
        """Escribe una transaccion y le aplica el color que le corresponde.

        Args:
            hoja: Hoja destino.
            fila: Numero de fila (1 es el encabezado).
            transaccion: Transaccion a volcar.
        """
        relleno = self._relleno_de(transaccion)

        for numero, columna in enumerate(self.columnas, start=1):
            celda = hoja.cell(row=fila, column=numero, value=columna.valor(transaccion))
            if columna.formato:
                celda.number_format = columna.formato
            if relleno is not None:
                celda.fill = relleno

    @staticmethod
    def _relleno_de(transaccion: TransaccionReconciliada) -> Optional[PatternFill]:
        """Elige el color de la fila segun la precedencia definida.

        Primero naranja si hay fraude; si no, rojo ante cualquier discrepancia
        o faltante; si no, verde para lo reconciliado.

        Args:
            transaccion: Transaccion a colorear.

        Returns:
            El relleno correspondiente, o `None` si no aplica ninguno.
        """
        if transaccion.es_fraude:
            return RELLENO_FRAUDE
        if not transaccion.esta_reconciliada:
            return RELLENO_HALLAZGO
        return RELLENO_RECONCILIADO

    def _dar_formato(self, hoja: Worksheet, filas: int) -> None:
        """Aplica congelado, autofiltro y ancho de columna.

        Args:
            hoja: Hoja a formatear.
            filas: Numero total de filas escritas, encabezado incluido.
        """
        hoja.freeze_panes = "A2"
        hoja.auto_filter.ref = f"A1:{get_column_letter(len(self.columnas))}{filas}"

        for numero, columna in enumerate(self.columnas, start=1):
            letra = get_column_letter(numero)
            ancho = max(
                len(columna.titulo) + 4,
                self._ancho_contenido(hoja, numero, filas),
            )
            hoja.column_dimensions[letra].width = min(max(ancho, ANCHO_MINIMO), ANCHO_MAXIMO)

    @staticmethod
    def _ancho_contenido(hoja: Worksheet, columna: int, filas: int) -> int:
        """Calcula el ancho necesario para el contenido de una columna.

        Args:
            hoja: Hoja donde medir.
            columna: Numero de columna.
            filas: Ultima fila con datos.

        Returns:
            El ancho en caracteres del valor mas largo de la columna.
        """
        maximo = 0
        for fila in range(2, filas + 1):
            valor = hoja.cell(row=fila, column=columna).value
            if valor is not None:
                maximo = max(maximo, len(str(valor)))
        return maximo + 2
